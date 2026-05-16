# -*- coding: utf-8 -*-
"""Excel导出模块测试"""

import os
import tempfile
import pytest
import gpxpy.gpx
from openpyxl import load_workbook
from gpx_editor.core.excel_exporter import ExcelExporter


def _make_waypoint(name, lat, lon, ele=None, desc=None, time=None):
    """创建测试用航点"""
    wp = gpxpy.gpx.GPXWaypoint(latitude=lat, longitude=lon, elevation=ele, name=name, description=desc)
    if time:
        from datetime import datetime
        wp.time = datetime(2026, 5, 16, 10, 30, 0)
    return wp


class TestExcelExporter:
    """ExcelExporter 测试"""

    def test_export_all_fields(self):
        """测试导出全部字段"""
        waypoints = [_make_waypoint("航点A", 39.9, 116.3, ele=100.5, desc="测试描述")]
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            fields = ["name", "latitude", "longitude", "elevation", "description", "time"]
            result = ExcelExporter.export(waypoints, fields, path)
            assert result is True
            wb = load_workbook(path)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            assert headers == ["名称", "纬度", "经度", "海拔", "描述", "时间"]
            assert ws.cell(2, 1).value == "航点A"
            assert ws.cell(2, 2).value == 39.9
            assert ws.cell(2, 3).value == 116.3
            assert ws.cell(2, 4).value == 100.5
            assert ws.cell(2, 5).value == "测试描述"
            wb.close()
        finally:
            os.unlink(path)

    def test_export_selected_fields(self):
        """测试导出部分字段"""
        waypoints = [_make_waypoint("航点A", 39.9, 116.3)]
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            fields = ["name", "latitude", "longitude"]
            result = ExcelExporter.export(waypoints, fields, path)
            assert result is True
            wb = load_workbook(path)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            assert headers == ["名称", "纬度", "经度"]
            assert ws.cell(2, 1).value == "航点A"
            wb.close()
        finally:
            os.unlink(path)

    def test_export_empty_values(self):
        """测试空值处理"""
        waypoints = [_make_waypoint("航点A", 39.9, 116.3)]
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            fields = ["name", "elevation", "description"]
            result = ExcelExporter.export(waypoints, fields, path)
            assert result is True
            wb = load_workbook(path)
            ws = wb.active
            assert ws.cell(2, 1).value == "航点A"
            assert ws.cell(2, 2).value is None
            assert ws.cell(2, 3).value is None
            wb.close()
        finally:
            os.unlink(path)

    def test_export_multiple_waypoints(self):
        """测试多条航点导出"""
        waypoints = [
            _make_waypoint("航点A", 39.9, 116.3),
            _make_waypoint("航点B", 40.0, 116.4),
            _make_waypoint("航点C", 40.1, 116.5),
        ]
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            fields = ["name", "latitude", "longitude"]
            result = ExcelExporter.export(waypoints, fields, path)
            assert result is True
            wb = load_workbook(path)
            ws = wb.active
            assert ws.max_row == 4
            assert ws.cell(2, 1).value == "航点A"
            assert ws.cell(3, 1).value == "航点B"
            assert ws.cell(4, 1).value == "航点C"
            wb.close()
        finally:
            os.unlink(path)

    def test_export_empty_waypoints(self):
        """测试空列表"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            fields = ["name"]
            result = ExcelExporter.export([], fields, path)
            assert result is True
            wb = load_workbook(path)
            ws = wb.active
            assert ws.max_row == 1
            wb.close()
        finally:
            os.unlink(path)
