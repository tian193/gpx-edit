# -*- coding: utf-8 -*-
"""
Excel导出模块
功能: 将航点数据导出为Excel格式
"""

import os
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


# 字段配置：代码名 -> 中文表头
FIELD_CONFIG = {
    "source_file": "来源文件",
    "name": "名称",
    "latitude": "纬度",
    "longitude": "经度",
    "elevation": "海拔",
    "description": "描述",
    "time": "时间",
}


def _get_waypoint_value(waypoint, field: str, source_file: str = None):
    """从航点对象获取指定字段值"""
    if field == "source_file":
        if source_file:
            return os.path.basename(source_file)
        return None
    if field == "time":
        if waypoint.time:
            return waypoint.time.strftime("%Y-%m-%d %H:%M:%S")
        return None
    return getattr(waypoint, field, None)


def _get_display_width(text: str) -> int:
    """计算字符串显示宽度，中文字符算2个宽度"""
    width = 0
    for char in text:
        if '一' <= char <= '鿿' or '　' <= char <= '〿':
            width += 2
        else:
            width += 1
    return width


class ExcelExporter:
    """Excel导出器"""

    @staticmethod
    def export(waypoints: list, selected_fields: List[str], output_path: str, source_files: dict = None):
        """
        导出航点到Excel
        Args:
            waypoints: gpxpy GPXWaypoint列表
            selected_fields: 选中的字段代码列表
            output_path: 输出文件路径(.xlsx)
            source_files: 航点来源文件映射 {waypoint_id: file_path}，可选
        """
        wb = Workbook()
        try:
            ws = wb.active
            ws.title = "航点数据"

            # 写表头
            headers = [FIELD_CONFIG[f] for f in selected_fields if f in FIELD_CONFIG]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

            # 写数据
            for row_idx, wp in enumerate(waypoints, 2):
                for col_idx, field in enumerate(selected_fields, 1):
                    if field in FIELD_CONFIG:
                        source_file = source_files.get(id(wp)) if source_files else None
                        value = _get_waypoint_value(wp, field, source_file)
                        ws.cell(row=row_idx, column=col_idx, value=value)

            # 自动调整列宽（考虑中文字符宽度）
            for col_idx, field in enumerate(selected_fields, 1):
                if field in FIELD_CONFIG:
                    max_width = _get_display_width(FIELD_CONFIG[field])
                    for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                        for cell in row:
                            if cell.value:
                                max_width = max(max_width, _get_display_width(str(cell.value)))
                    ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = min(max_width + 4, 50)

            wb.save(output_path)
        finally:
            wb.close()
