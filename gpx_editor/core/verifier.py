# -*- coding: utf-8 -*-
"""
三方数据一致性校验模块
功能: 对比分配航点、GPS实际航点、手写样品编号三者是否一致
"""

import os
import csv
from enum import Enum
from typing import Dict, List, Optional, Callable
from dataclasses import dataclass, field

import gpxpy
import gpxpy.gpx

from .batch_matcher import BatchMatcher


class MatchStatus(Enum):
    """匹配状态"""
    MATCHED = "一致"
    MISSING_GPS = "GPS未到达"
    EXTRA_GPS = "GPS多余"
    MISSING_SAMPLE = "缺少样品"
    EXTRA_SAMPLE = "多余样品"
    DISTANCE_WARNING = "距离偏差"


@dataclass
class WaypointRecord:
    """单个航点记录（从任一数据源提取）"""
    name: str
    latitude: Optional[float] = None
    longitude: Optional[float] = None
    elevation: Optional[float] = None
    source_file: str = ""


@dataclass
class SampleRecord:
    """样品记录（从Excel/CSV提取）"""
    waypoint_name: str
    sample_id: str
    row_number: int = 0
    extra_fields: Dict[str, str] = field(default_factory=dict)


@dataclass
class VerificationItem:
    """单条校验结果"""
    waypoint_name: str
    assigned: Optional[WaypointRecord] = None
    gps_actual: Optional[WaypointRecord] = None
    sample: Optional[SampleRecord] = None
    distance: Optional[float] = None
    status: MatchStatus = MatchStatus.MATCHED
    notes: str = ""


@dataclass
class VerificationResult:
    """校验结果汇总"""
    group_label: str
    assigned_file: str
    gps_file: str
    sample_file: str
    items: List[VerificationItem] = field(default_factory=list)

    @property
    def total_assigned(self) -> int:
        return sum(1 for i in self.items if i.assigned is not None)

    @property
    def total_gps(self) -> int:
        return sum(1 for i in self.items if i.gps_actual is not None)

    @property
    def total_samples(self) -> int:
        return sum(1 for i in self.items if i.sample is not None)

    @property
    def total_matched(self) -> int:
        return sum(1 for i in self.items if i.status == MatchStatus.MATCHED)

    @property
    def total_issues(self) -> int:
        return sum(1 for i in self.items if i.status != MatchStatus.MATCHED)


@dataclass
class BatchVerificationResult:
    """批量校验结果"""
    group_results: List[VerificationResult] = field(default_factory=list)
    errors: List[str] = field(default_factory=list)


# 文件名匹配模式
ASSIGNED_PATTERNS = ["分配", "assigned", "baseline", "计划"]
GPS_PATTERNS = ["gps", "实际", "actual", "记录"]
SAMPLE_PATTERNS = ["样品", "sample", "编号", "样本"]


class Verifier:
    """三方数据一致性校验器"""

    DISTANCE_THRESHOLD = 200.0  # 默认距离阈值(米)

    @staticmethod
    def load_gpx_waypoints(gpx_path: str) -> List[WaypointRecord]:
        """加载GPX文件中的航点"""
        records = []
        try:
            with open(gpx_path, 'r', encoding='utf-8') as f:
                gpx = gpxpy.parse(f)
        except UnicodeDecodeError:
            with open(gpx_path, 'r', encoding='gbk') as f:
                gpx = gpxpy.parse(f)

        for wpt in gpx.waypoints:
            records.append(WaypointRecord(
                name=wpt.name or "",
                latitude=wpt.latitude,
                longitude=wpt.longitude,
                elevation=wpt.elevation,
                source_file=os.path.basename(gpx_path)
            ))
        return records

    @staticmethod
    def load_samples(file_path: str) -> List[SampleRecord]:
        """加载样品编号文件（支持 .xlsx / .csv）"""
        ext = os.path.splitext(file_path)[1].lower()

        if ext == '.xlsx':
            return Verifier._load_samples_xlsx(file_path)
        elif ext == '.csv':
            return Verifier._load_samples_csv(file_path)
        else:
            raise ValueError(f"不支持的文件格式: {ext}")

    @staticmethod
    def _load_samples_xlsx(file_path: str) -> List[SampleRecord]:
        """从Excel文件加载样品数据"""
        from openpyxl import load_workbook

        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []

        headers = [str(h).strip() if h else "" for h in rows[0]]
        name_col, sample_col = Verifier._detect_columns(headers)

        if name_col is None or sample_col is None:
            raise ValueError("无法自动识别航点名称列和样品编号列，请检查表头")

        samples = []
        for row_idx, row in enumerate(rows[1:], start=2):
            name_val = row[name_col] if name_col < len(row) else None
            sample_val = row[sample_col] if sample_col < len(row) else None

            if name_val is None or sample_val is None:
                continue

            name = str(name_val).strip()
            sample_id = str(sample_val).strip()

            if not name or not sample_id:
                continue

            extra = {}
            for i, h in enumerate(headers):
                if i != name_col and i != sample_col and i < len(row) and row[i]:
                    extra[h] = str(row[i])

            samples.append(SampleRecord(
                waypoint_name=name,
                sample_id=sample_id,
                row_number=row_idx,
                extra_fields=extra
            ))

        wb.close()
        return samples

    @staticmethod
    def _load_samples_csv(file_path: str) -> List[SampleRecord]:
        """从CSV文件加载样品数据"""
        # 尝试多种编码
        for encoding in ['utf-8', 'gbk', 'gb2312', 'latin-1']:
            try:
                return Verifier._parse_csv(file_path, encoding)
            except (UnicodeDecodeError, UnicodeError):
                continue
        raise ValueError(f"无法读取CSV文件: {file_path}")

    @staticmethod
    def _parse_csv(file_path: str, encoding: str) -> List[SampleRecord]:
        """解析CSV文件"""
        samples = []
        with open(file_path, 'r', encoding=encoding) as f:
            reader = csv.DictReader(f)
            headers = reader.fieldnames or []

            name_key, sample_key = Verifier._detect_csv_columns(headers)

            if name_key is None or sample_key is None:
                raise ValueError("无法自动识别航点名称列和样品编号列，请检查表头")

            for row_idx, row in enumerate(reader, start=2):
                name = row.get(name_key, "").strip()
                sample_id = row.get(sample_key, "").strip()

                if not name or not sample_id:
                    continue

                extra = {k: v for k, v in row.items()
                         if k != name_key and k != sample_key and v}

                samples.append(SampleRecord(
                    waypoint_name=name,
                    sample_id=sample_id,
                    row_number=row_idx,
                    extra_fields=extra
                ))

        return samples

    @staticmethod
    def _detect_columns(headers: List[str]):
        """检测Excel表头中的航点名称列和样品编号列"""
        name_keywords = ["航点", "名称", "点名", "name", "waypoint", "wpt", "点号"]
        sample_keywords = ["样品", "编号", "样本", "sample", "id", "样号"]

        name_col = None
        sample_col = None

        for i, h in enumerate(headers):
            h_lower = h.lower()
            if name_col is None:
                for kw in name_keywords:
                    if kw in h_lower:
                        name_col = i
                        break
            if sample_col is None:
                for kw in sample_keywords:
                    if kw in h_lower:
                        sample_col = i
                        break

        return name_col, sample_col

    @staticmethod
    def _detect_csv_columns(headers: List[str]):
        """检测CSV表头中的航点名称列和样品编号列"""
        name_keywords = ["航点", "名称", "点名", "name", "waypoint", "wpt", "点号"]
        sample_keywords = ["样品", "编号", "样本", "sample", "id", "样号"]

        name_key = None
        sample_key = None

        for h in headers:
            h_lower = h.lower()
            if name_key is None:
                for kw in name_keywords:
                    if kw in h_lower:
                        name_key = h
                        break
            if sample_key is None:
                for kw in sample_keywords:
                    if kw in h_lower:
                        sample_key = h
                        break

        return name_key, sample_key

    @staticmethod
    def verify_single(
        assigned: List[WaypointRecord],
        gps: List[WaypointRecord],
        samples: List[SampleRecord],
        distance_threshold: float = 200.0
    ) -> List[VerificationItem]:
        """单次校验：三方对比"""
        # 按名称建立索引（不区分大小写，去空格）
        assigned_map = {w.name.strip().lower(): w for w in assigned if w.name}
        gps_map = {w.name.strip().lower(): w for w in gps if w.name}
        sample_map = {s.waypoint_name.strip().lower(): s for s in samples if s.waypoint_name}

        # 取所有名称的并集
        all_names = set(assigned_map.keys()) | set(gps_map.keys()) | set(sample_map.keys())

        items = []
        for name_key in all_names:
            a = assigned_map.get(name_key)
            g = gps_map.get(name_key)
            s = sample_map.get(name_key)

            # 显示名称优先使用分配航点的名称
            if a:
                display_name = a.name
            elif g:
                display_name = g.name
            elif s:
                display_name = s.waypoint_name
            else:
                display_name = name_key

            item = VerificationItem(
                waypoint_name=display_name,
                assigned=a,
                gps_actual=g,
                sample=s
            )

            # 计算距离
            if a and g and a.latitude and g.latitude:
                item.distance = BatchMatcher.calculate_distance(
                    a.latitude, a.longitude, g.latitude, g.longitude
                )

            # 判断状态
            if a and g and s:
                if item.distance and item.distance > distance_threshold:
                    item.status = MatchStatus.DISTANCE_WARNING
                    item.notes = f"距离偏差 {item.distance:.0f}米"
                else:
                    item.status = MatchStatus.MATCHED
            elif a and g and not s:
                item.status = MatchStatus.MISSING_SAMPLE
                item.notes = "有分配和GPS记录，但缺少样品编号"
            elif a and s and not g:
                item.status = MatchStatus.MISSING_GPS
                item.notes = "有分配和样品，但GPS未到达"
            elif g and s and not a:
                item.status = MatchStatus.EXTRA_GPS
                item.notes = "GPS和样品存在，但不在分配列表中"
            elif a and not g and not s:
                item.status = MatchStatus.MISSING_GPS
                item.notes = "仅有分配，GPS和样品均缺失"
            elif g and not a and not s:
                item.status = MatchStatus.EXTRA_GPS
                item.notes = "仅有GPS记录，不在分配列表中"
            elif s and not a and not g:
                item.status = MatchStatus.EXTRA_SAMPLE
                item.notes = "仅有样品记录，不在分配列表中"

            items.append(item)

        # 排序：问题项在前，按状态优先级排序
        status_order = {
            MatchStatus.MISSING_GPS: 0,
            MatchStatus.EXTRA_GPS: 1,
            MatchStatus.MISSING_SAMPLE: 2,
            MatchStatus.EXTRA_SAMPLE: 3,
            MatchStatus.DISTANCE_WARNING: 4,
            MatchStatus.MATCHED: 5
        }
        items.sort(key=lambda x: (status_order.get(x.status, 99), x.waypoint_name))

        return items

    @staticmethod
    def scan_directory(base_dir: str) -> List[Dict[str, str]]:
        """扫描目录，识别分组和文件"""
        groups = []

        # 检查是否有子目录结构
        subdirs = [d for d in os.listdir(base_dir)
                   if os.path.isdir(os.path.join(base_dir, d))]

        if subdirs:
            # 子目录结构：日期/小组/文件
            for subdir in subdirs:
                subdir_path = os.path.join(base_dir, subdir)
                group = Verifier._scan_group_dir(subdir_path)
                if group:
                    group['label'] = subdir
                    groups.append(group)

                # 检查是否有二级子目录（日期/小组/文件）
                sub_subdirs = [d for d in os.listdir(subdir_path)
                               if os.path.isdir(os.path.join(subdir_path, d))]
                for sub_subdir in sub_subdirs:
                    sub_subdir_path = os.path.join(subdir_path, sub_subdir)
                    group = Verifier._scan_group_dir(sub_subdir_path)
                    if group:
                        group['label'] = f"{subdir}/{sub_subdir}"
                        groups.append(group)
        else:
            # 平铺结构：按文件名模式分组
            group = Verifier._scan_group_dir(base_dir)
            if group:
                group['label'] = os.path.basename(base_dir)
                groups.append(group)

        return groups

    @staticmethod
    def _scan_group_dir(dir_path: str) -> Optional[Dict[str, str]]:
        """扫描单个目录，识别三类文件"""
        result = {'assigned': None, 'gps': None, 'sample': None}

        for filename in os.listdir(dir_path):
            filepath = os.path.join(dir_path, filename)
            if not os.path.isfile(filepath):
                continue

            ext = os.path.splitext(filename)[1].lower()
            name_lower = filename.lower()

            # 识别分配航点GPX
            if ext == '.gpx' and any(p in name_lower for p in ASSIGNED_PATTERNS):
                result['assigned'] = filepath

            # 识别GPS实际GPX
            elif ext == '.gpx' and any(p in name_lower for p in GPS_PATTERNS):
                result['gps'] = filepath

            # 识别样品文件
            elif ext in ('.xlsx', '.csv') and any(p in name_lower for p in SAMPLE_PATTERNS):
                result['sample'] = filepath

            # 兜底：未标记的GPX文件
            elif ext == '.gpx':
                if result['assigned'] is None:
                    result['assigned'] = filepath
                elif result['gps'] is None:
                    result['gps'] = filepath

            # 兜底：未标记的Excel/CSV文件
            elif ext in ('.xlsx', '.csv'):
                if result['sample'] is None:
                    result['sample'] = filepath

        # 至少需要有分配航点和GPS航点
        if result['assigned'] and result['gps']:
            return result
        return None

    @staticmethod
    def verify_batch(
        base_dir: str,
        distance_threshold: float = 200.0,
        callback: Optional[Callable] = None
    ) -> BatchVerificationResult:
        """批量校验：按目录结构自动分组"""
        batch_result = BatchVerificationResult()
        groups = Verifier.scan_directory(base_dir)

        total = len(groups)
        for idx, group in enumerate(groups):
            try:
                if callback:
                    callback(group['label'], idx + 1, total)

                assigned = Verifier.load_gpx_waypoints(group['assigned'])
                gps = Verifier.load_gpx_waypoints(group['gps'])
                samples = []

                if group['sample']:
                    samples = Verifier.load_samples(group['sample'])

                items = Verifier.verify_single(assigned, gps, samples, distance_threshold)

                result = VerificationResult(
                    group_label=group['label'],
                    assigned_file=os.path.basename(group['assigned']),
                    gps_file=os.path.basename(group['gps']),
                    sample_file=os.path.basename(group['sample']) if group['sample'] else "",
                    items=items
                )
                batch_result.group_results.append(result)

            except Exception as e:
                batch_result.errors.append(f"{group['label']}: {str(e)}")

        return batch_result

    @staticmethod
    def export_to_excel(results: List[VerificationResult], output_path: str):
        """导出校验结果到Excel"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()

        # 颜色定义
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        blue_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        status_fills = {
            MatchStatus.MATCHED: green_fill,
            MatchStatus.MISSING_GPS: red_fill,
            MatchStatus.EXTRA_GPS: blue_fill,
            MatchStatus.MISSING_SAMPLE: yellow_fill,
            MatchStatus.EXTRA_SAMPLE: red_fill,
            MatchStatus.DISTANCE_WARNING: yellow_fill
        }

        # Sheet 1: 汇总
        ws_summary = wb.active
        ws_summary.title = "汇总"
        summary_headers = ["分组", "分配航点数", "GPS航点数", "样品数", "一致数", "问题数"]
        for col, header in enumerate(summary_headers, 1):
            cell = ws_summary.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row_idx, result in enumerate(results, 2):
            ws_summary.cell(row=row_idx, column=1, value=result.group_label)
            ws_summary.cell(row=row_idx, column=2, value=result.total_assigned)
            ws_summary.cell(row=row_idx, column=3, value=result.total_gps)
            ws_summary.cell(row=row_idx, column=4, value=result.total_samples)
            ws_summary.cell(row=row_idx, column=5, value=result.total_matched)
            ws_summary.cell(row=row_idx, column=6, value=result.total_issues)

            # 问题数标红
            if result.total_issues > 0:
                ws_summary.cell(row=row_idx, column=6).fill = red_fill

        # 调整列宽
        ws_summary.column_dimensions['A'].width = 20
        for col in 'BCDEF':
            ws_summary.column_dimensions[col].width = 14

        # Sheet 2+: 每个分组的详细结果
        for result in results:
            sheet_name = result.group_label[:31]
            ws = wb.create_sheet(title=sheet_name)

            detail_headers = ["航点名称", "分配坐标", "GPS坐标", "样品编号", "距离(米)", "状态", "备注"]
            for col, header in enumerate(detail_headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            for row_idx, item in enumerate(result.items, 2):
                # 航点名称
                ws.cell(row=row_idx, column=1, value=item.waypoint_name)

                # 分配坐标
                if item.assigned and item.assigned.latitude:
                    coord = f"{item.assigned.latitude:.6f}, {item.assigned.longitude:.6f}"
                    ws.cell(row=row_idx, column=2, value=coord)
                else:
                    ws.cell(row=row_idx, column=2, value="(无)")

                # GPS坐标
                if item.gps_actual and item.gps_actual.latitude:
                    coord = f"{item.gps_actual.latitude:.6f}, {item.gps_actual.longitude:.6f}"
                    ws.cell(row=row_idx, column=3, value=coord)
                else:
                    ws.cell(row=row_idx, column=3, value="(无)")

                # 样品编号
                if item.sample:
                    ws.cell(row=row_idx, column=4, value=item.sample.sample_id)
                else:
                    ws.cell(row=row_idx, column=4, value="(无)")

                # 距离
                if item.distance is not None:
                    ws.cell(row=row_idx, column=5, value=f"{item.distance:.1f}")
                else:
                    ws.cell(row=row_idx, column=5, value="-")

                # 状态
                ws.cell(row=row_idx, column=6, value=item.status.value)

                # 备注
                ws.cell(row=row_idx, column=7, value=item.notes)

                # 应用颜色
                fill = status_fills.get(item.status)
                if fill:
                    for col in range(1, 8):
                        ws.cell(row=row_idx, column=col).fill = fill

            # 调整列宽
            ws.column_dimensions['A'].width = 16
            ws.column_dimensions['B'].width = 22
            ws.column_dimensions['C'].width = 22
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 12
            ws.column_dimensions['G'].width = 30

        wb.save(output_path)
